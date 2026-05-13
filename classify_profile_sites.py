from __future__ import annotations

import argparse
import random
import re
import socket
import time
from dataclasses import dataclass
from html import unescape
from pathlib import Path
from typing import Callable, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urljoin, urlparse
from urllib.request import Request, build_opener

import openpyxl
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By


WORKBOOK_PATH = Path("Large Free Profile Submission Site List.xlsx")
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)
KEYWORDS = [
    ("Login", "login"),
    ("Sign In", "sign in"),
    ("Register", "register"),
    ("Sign Up", "sign up"),
]
AUTH_PATHS = {
    "Login": ["/login", "/signin", "/sign-in"],
    "Sign In": ["/signin", "/sign-in", "/login"],
    "Register": ["/register", "/signup", "/sign-up", "/join"],
    "Sign Up": ["/signup", "/sign-up", "/register", "/join"],
}
TYPE_PATTERNS = [
    ("Forum", [r"\bforum\b", r"\bcommunity\b", r"\bdiscussion\b", r"\bthread\b"]),
    ("Social Network", [r"\bsocial\b", r"\bfollow\b", r"\bprofile\b", r"\bshare\b"]),
    ("Blog", [r"\bblog\b", r"\bpost\b", r"\barticle\b", r"\bcomments\b"]),
    ("Business Directory", [r"\bdirectory\b", r"\blisting\b", r"\bbusiness\b", r"\blocal\b"]),
    ("Ecommerce", [r"\bshop\b", r"\bstore\b", r"\bcart\b", r"\bproduct\b"]),
    ("News/Media", [r"\bnews\b", r"\bmedia\b", r"\bmagazine\b", r"\bpress\b"]),
    ("Education", [r"\blearn\b", r"\bcourse\b", r"\beducation\b", r"\bstudent\b"]),
    ("Portfolio/Creative", [r"\bportfolio\b", r"\bgallery\b", r"\bcreative\b", r"\bdesign\b"]),
    ("Wiki/Knowledge", [r"\bwiki\b", r"\bknowledge\b", r"\bdocs\b", r"\bdocumentation\b"]),
    ("Job Board", [r"\bjobs?\b", r"\bcareer\b", r"\bhiring\b", r"\bresume\b"]),
    ("Real Estate", [r"\breal estate\b", r"\bproperty\b", r"\bhomes?\b", r"\brealtor\b"]),
]
REQUEST_TIMEOUT = 12
SAVE_EVERY = 25


@dataclass
class FetchResult:
    final_url: str | None
    html: str
    visible_text: str
    status: str
    engine: str = ""


def normalize_url(value: str) -> list[str]:
    raw = (value or "").strip()
    if not raw:
        return []
    raw = raw.split("#", 1)[0].strip()
    candidates = []
    if raw.startswith(("http://", "https://")):
        candidates.append(raw)
    else:
        candidates.extend([f"https://{raw}", f"http://{raw}"])
    return candidates


def get_opener():
    return build_opener()


def fetch_requests(urls: Iterable[str]) -> FetchResult:
    opener = get_opener()
    last_error = "Fetch failed"
    for url in urls:
        req = Request(
            url,
            headers={
                "User-Agent": USER_AGENT,
                "Accept": "text/html,application/xhtml+xml",
                "Accept-Language": "en-US,en;q=0.9",
            },
        )
        try:
            with opener.open(req, timeout=REQUEST_TIMEOUT) as resp:
                content_type = resp.headers.get("Content-Type", "")
                body = resp.read(1024 * 1024)
                text = body.decode("utf-8", errors="ignore")
                visible_text = strip_html(text)
                if "text/html" not in content_type.lower() and "<html" not in text.lower():
                    return FetchResult(str(resp.geturl()), text, visible_text, "Non-HTML page", "requests")
                return FetchResult(str(resp.geturl()), text, visible_text, "OK", "requests")
        except HTTPError as exc:
            last_error = f"HTTP {exc.code}"
        except URLError as exc:
            reason = getattr(exc, "reason", exc)
            last_error = f"URL error: {reason}"
        except socket.timeout:
            last_error = "Timeout"
        except Exception as exc:  # noqa: BLE001
            last_error = f"Error: {exc}"
    return FetchResult(None, "", "", last_error, "requests")


def create_webdriver(headless: bool) -> webdriver.Chrome:
    options = Options()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1440,2200")
    options.add_argument("--log-level=3")
    options.add_argument(f"--user-agent={USER_AGENT}")
    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(REQUEST_TIMEOUT)
    return driver


def fetch_selenium(
    urls: Iterable[str],
    driver: webdriver.Chrome,
    render_wait: float,
) -> FetchResult:
    last_error = "Fetch failed"
    for url in urls:
        try:
            driver.get(url)
            if render_wait > 0:
                time.sleep(render_wait)
            html = driver.page_source
            visible_text = driver.find_element(By.TAG_NAME, "body").text.strip()
            return FetchResult(driver.current_url, html, visible_text, "OK", "selenium")
        except TimeoutException:
            last_error = "Timeout"
        except WebDriverException as exc:
            last_error = f"Browser error: {exc.msg.splitlines()[0]}"
        except Exception as exc:  # noqa: BLE001
            last_error = f"Error: {exc}"
    return FetchResult(None, "", "", last_error, "selenium")


def strip_html(html: str) -> str:
    text = re.sub(r"(?is)<script.*?>.*?</script>", " ", html)
    text = re.sub(r"(?is)<style.*?>.*?</style>", " ", text)
    text = re.sub(r"(?s)<[^>]+>", " ", text)
    text = unescape(text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def detect_keywords_from_text(visible_text: str, final_url: str | None) -> list[str]:
    lower_text = visible_text.lower()
    matches: list[str] = []
    for label, needle in KEYWORDS:
        if needle in lower_text:
            matches.append(label)

    if final_url:
        parsed = urlparse(final_url)
        path = parsed.path.lower()
        for label, patterns in AUTH_PATHS.items():
            if label not in matches and any(part in path for part in patterns):
                matches.append(label)

    return matches


def base_site_urls(value: str) -> list[str]:
    candidates = normalize_url(value)
    base_urls: list[str] = []
    for candidate in candidates:
        parsed = urlparse(candidate)
        if not parsed.scheme or not parsed.netloc:
            continue
        base_urls.append(f"{parsed.scheme}://{parsed.netloc}/")
    return list(dict.fromkeys(base_urls))


def probe_auth_paths(
    value: str,
    found_labels: list[str],
    fetcher,
) -> list[str]:
    matches = list(found_labels)
    base_urls = base_site_urls(value)
    for label, paths in AUTH_PATHS.items():
        if label in matches:
            continue
        for base_url in base_urls:
            for path in paths:
                result = fetcher([urljoin(base_url, path)])
                if result.status == "OK":
                    text_matches = detect_keywords_from_text(result.visible_text, result.final_url)
                    if label in text_matches:
                        matches.append(label)
                        break
                    parsed = urlparse(result.final_url or "")
                    if any(auth_path in parsed.path.lower() for auth_path in paths):
                        matches.append(label)
                        break
            if label in matches:
                break
    return matches


def should_flag_js_heavy(html: str, visible_text: str) -> bool:
    if visible_text:
        return False
    html_lower = html.lower()
    markers = [
        "__next",
        "id=\"app\"",
        "id='app'",
        "id=\"root\"",
        "id='root'",
        "ng-app",
        "data-reactroot",
        "application/ld+json",
    ]
    return any(marker in html_lower for marker in markers)


def should_retry_with_selenium(result: FetchResult, js_mode: str) -> bool:
    if result.status != "OK":
        return False
    if detect_keywords_from_text(result.visible_text, result.final_url):
        return False
    if should_flag_js_heavy(result.html, result.visible_text):
        return True
    if js_mode != "off" and len(result.visible_text.strip()) < 80:
        return True
    return False


def detect_keywords(
    homepage_result: FetchResult,
    source_value: str,
    scan_auth_paths: bool,
    js_mode: str,
    fetcher,
) -> tuple[str, str]:
    matches = detect_keywords_from_text(homepage_result.visible_text, homepage_result.final_url)
    note_parts: list[str] = []

    if scan_auth_paths:
        before = set(matches)
        matches = probe_auth_paths(source_value, matches, fetcher)
        new_matches = [label for label in matches if label not in before]
        if new_matches:
            note_parts.append(f"auth-path match: {', '.join(new_matches)}")

    if js_mode != "off" and should_flag_js_heavy(homepage_result.html, homepage_result.visible_text):
        note_parts.append("JS-heavy site detected")

    status_text = ", ".join(matches) if matches else "Not found"
    note_text = " | ".join(note_parts)
    return status_text, note_text


def build_fetchers(
    fetch_mode: str,
    render_wait: float,
) -> tuple[Callable[[Iterable[str]], FetchResult], Callable[[Iterable[str]], FetchResult] | None, webdriver.Chrome | None]:
    driver = None
    selenium_fetcher = None

    if fetch_mode in {"selenium", "hybrid"}:
        driver = create_webdriver(headless=True)
        selenium_fetcher = lambda urls: fetch_selenium(urls, driver, render_wait)

    if fetch_mode == "selenium":
        primary_fetcher = selenium_fetcher
    else:
        primary_fetcher = fetch_requests

    return primary_fetcher, selenium_fetcher, driver


def ask_yes_no(prompt: str, default: bool) -> bool:
    suffix = "Y/n" if default else "y/N"
    raw = input(f"{prompt} [{suffix}]: ").strip().lower()
    if not raw:
        return default
    return raw in {"y", "yes"}


def ask_int(prompt: str, default: int, minimum: int = 0) -> int:
    raw = input(f"{prompt} [{default}]: ").strip()
    if not raw:
        return default
    try:
        value = int(raw)
    except ValueError:
        print(f"Invalid number, using {default}.")
        return default
    if value < minimum:
        print(f"Value must be at least {minimum}, using {default}.")
        return default
    return value


def ask_float(prompt: str, default: float, minimum: float = 0.0) -> float:
    raw = input(f"{prompt} [{default}]: ").strip()
    if not raw:
        return default
    try:
        value = float(raw)
    except ValueError:
        print(f"Invalid number, using {default}.")
        return default
    if value < minimum:
        print(f"Value must be at least {minimum}, using {default}.")
        return default
    return value


def ask_choice(prompt: str, choices: list[str], default: str) -> str:
    print(prompt)
    for index, choice in enumerate(choices, start=1):
        default_tag = " (default)" if choice == default else ""
        print(f"  {index}. {choice}{default_tag}")
    raw = input("Choose option: ").strip()
    if not raw:
        return default
    if raw.isdigit():
        index = int(raw) - 1
        if 0 <= index < len(choices):
            return choices[index]
    lowered = raw.lower()
    for choice in choices:
        if choice.lower() == lowered:
            return choice
    print(f"Invalid choice, using {default}.")
    return default


def apply_interactive_menu(args: argparse.Namespace) -> argparse.Namespace:
    print("Interactive run menu")
    print("--------------------")
    args.fetch_mode = ask_choice(
        "Fetch mode:",
        ["hybrid", "requests", "selenium"],
        args.fetch_mode,
    )
    args.scan_auth_paths = ask_yes_no("Scan common auth paths like /login and /signup?", args.scan_auth_paths)
    args.stop_if_locked = ask_yes_no("Stop immediately if the original workbook is locked?", args.stop_if_locked)
    args.js_mode = ask_choice(
        "JS-heavy handling:",
        ["detect", "note", "off"],
        args.js_mode,
    )
    args.sample_size = ask_int("Sample size (0 = full run)", args.sample_size, minimum=0)
    if args.sample_size > 0:
        args.seed = ask_int("Random seed", args.seed, minimum=0)
    args.render_wait = ask_float("Selenium render wait in seconds", args.render_wait, minimum=0.0)
    print("")
    print("Selected options:")
    print(f"  fetch_mode={args.fetch_mode}")
    print(f"  scan_auth_paths={args.scan_auth_paths}")
    print(f"  stop_if_locked={args.stop_if_locked}")
    print(f"  js_mode={args.js_mode}")
    print(f"  sample_size={args.sample_size}")
    if args.sample_size > 0:
        print(f"  seed={args.seed}")
    print(f"  render_wait={args.render_wait}")
    print("")
    return args


def classify_site(url: str, html: str) -> str:
    corpus = " ".join(
        filter(
            None,
            [
                url.lower(),
                html[:100000].lower(),
            ],
        )
    )

    for label, patterns in TYPE_PATTERNS:
        for pattern in patterns:
            if re.search(pattern, corpus):
                return label
    return "General Website"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--stop-if-locked",
        action="store_true",
        help="Stop immediately if the original workbook cannot be saved.",
    )
    parser.add_argument(
        "--sample-size",
        type=int,
        default=0,
        help="Process a random sample of incomplete rows instead of the full sheet.",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=42,
        help="Random seed used when --sample-size is set.",
    )
    parser.add_argument(
        "--scan-auth-paths",
        action="store_true",
        help="Also probe common auth URLs like /login, /register, and /signup.",
    )
    parser.add_argument(
        "--js-mode",
        choices=["off", "detect", "note"],
        default="detect",
        help="How to handle JS-heavy sites: off=ignore, detect/note=mark likely JS-heavy pages.",
    )
    parser.add_argument(
        "--fetch-mode",
        choices=["selenium", "requests", "hybrid"],
        default="hybrid",
        help="Page loading engine. hybrid tries requests first, then Selenium for likely JS-heavy misses.",
    )
    parser.add_argument(
        "--render-wait",
        type=float,
        default=2.0,
        help="Seconds to wait after a page loads in Selenium mode for JS-rendered content.",
    )
    parser.add_argument(
        "--menu",
        action="store_true",
        help="Show an interactive menu for choosing run options.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Run checks and print results without saving anything to the workbook.",
    )
    parser.add_argument(
        "--last-n",
        type=int,
        default=0,
        help="Process only the last N incomplete rows.",
    )
    args = parser.parse_args()

    if args.menu:
        args = apply_interactive_menu(args)

    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb.active
    completed_rows = 0

    for row in range(1, ws.max_row + 1):
        existing_status = ws.cell(row=row, column=2).value
        existing_type = ws.cell(row=row, column=3).value
        if existing_status and existing_type:
            completed_rows = row
        else:
            break

    if completed_rows:
        print(f"Resuming after row {completed_rows}")
    else:
        print("Starting from row 1")

    rows_to_process = []
    for row in range(1, ws.max_row + 1):
        domain = ws.cell(row=row, column=1).value
        if not domain:
            continue

        existing_status = ws.cell(row=row, column=2).value
        existing_type = ws.cell(row=row, column=3).value
        if existing_status and existing_type:
            continue
        rows_to_process.append(row)

    if args.sample_size > 0:
        sample_count = min(args.sample_size, len(rows_to_process))
        rng = random.Random(args.seed)
        rows_to_process = sorted(rng.sample(rows_to_process, sample_count))
        print(f"Testing {sample_count} random rows: {rows_to_process}")
    elif args.last_n > 0:
        rows_to_process = rows_to_process[-args.last_n :]
        print(f"Processing last {len(rows_to_process)} incomplete rows: {rows_to_process}")

    driver = None
    try:
        fetcher, selenium_fetcher, driver = build_fetchers(args.fetch_mode, args.render_wait)

        for row in rows_to_process:
            domain = ws.cell(row=row, column=1).value

            result = fetcher(normalize_url(str(domain)))
            retried_with_selenium = False
            if (
                args.fetch_mode == "hybrid"
                and selenium_fetcher is not None
                and should_retry_with_selenium(result, args.js_mode)
            ):
                selenium_result = selenium_fetcher(normalize_url(str(domain)))
                if selenium_result.status == "OK":
                    result = selenium_result
                    retried_with_selenium = True

            status_text, note_text = detect_keywords(
                result,
                str(domain),
                scan_auth_paths=args.scan_auth_paths,
                js_mode=args.js_mode,
                fetcher=selenium_fetcher if retried_with_selenium and selenium_fetcher is not None else fetcher,
            )
            if result.status != "OK":
                status_text = f"{status_text} | {result.status}"
            if note_text:
                status_text = f"{status_text} | {note_text}"
            if retried_with_selenium:
                status_text = f"{status_text} | fetched with Selenium"

            site_type = classify_site(result.final_url or str(domain), result.visible_text or strip_html(result.html))

            if not args.dry_run:
                ws.cell(row=row, column=2, value=status_text)
                ws.cell(row=row, column=3, value=site_type)

            if not args.dry_run and row % SAVE_EVERY == 0:
                try:
                    wb.save(WORKBOOK_PATH)
                except PermissionError:
                    if args.stop_if_locked:
                        raise PermissionError(
                            f"Workbook is locked and --stop-if-locked is enabled: {WORKBOOK_PATH}"
                        )
                    raise
                print(f"Saved progress through row {row}")

            print(f"Row {row}: {domain} -> {status_text} | {site_type}")
            time.sleep(0.2)
    finally:
        if driver is not None:
            driver.quit()

    try:
        if not args.dry_run:
            wb.save(WORKBOOK_PATH)
    except PermissionError:
        if args.stop_if_locked:
            raise PermissionError(
                f"Workbook is locked and --stop-if-locked is enabled: {WORKBOOK_PATH}"
            )
        raise
    print("Done")


if __name__ == "__main__":
    main()
